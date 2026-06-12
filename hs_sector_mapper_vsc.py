"""
╔══════════════════════════════════════════════════════════════════════════════╗
║   IME Manufacturing Trade Mapper — VSCode / Local Edition                   ║
║   Converts commodity-wise trade xlsx (6-digit HS) → ISIC Rev 4 (2-digit)   ║
║   → IME NIC Sectors                                                         ║
║                                                                             ║
║   MULTI-YEAR SUPPORT: 2018-19 through 2024-25 (or any range)               ║
║   All years are combined into ONE output workbook with a master summary.    ║
║                                                                             ║
║   HOW TO USE                                                                ║
║   ─────────                                                                 ║
║   1. Place this file anywhere on your machine.                              ║
║   2. Create an "input" folder next to this script (auto-created on first    ║
║      run if missing).                                                       ║
║   3. Drop your xlsx files into the input folder. Supported naming:          ║
║        Trade volume : TradeStat_2018-19.xlsx  (or any name with "Trade")   ║
║        Partners     : Partners_2024-25.xlsx   (or any name with "Partner") ║
║      Any file containing "Trade" in its name is treated as a Tradestat      ║
║      file; any file containing "Partner" as a partners file.               ║
║   4. Run:  python hs_sector_mapper_vsc.py                                  ║
║   5. Output lands in an "output" folder next to this script.               ║
║                                                                             ║
║   OPTIONAL: Edit the CONFIG section below to customise paths, column        ║
║   names, and output file name.                                              ║
╚══════════════════════════════════════════════════════════════════════════════╝

NIC SECTOR MAPPING
──────────────────
  Food and Agro          : NIC 10, 11, 12
  Consumer Goods         : NIC 15
  Textiles               : NIC 13, 14
  Construction Materials : NIC 16, 23, 31
  Transport Equipment    : NIC 29, 30
  Engineering Goods      : NIC 28
  Electrical Equipment   : NIC 27
  Electronics            : NIC 26
  Metals                 : NIC 24, 25
  Chemicals              : NIC 19, 20, 21, 22
  General Manufacturing  : NIC 17, 18, 32

  MEMO ROW (not in total):
  Pharmaceuticals (memo) : NIC 21 — shown separately, NOT added to grand total
"""

# ═══════════════════════════════════════════════════════════════════════════
# CONFIGURATION  — edit these values to suit your setup
# ═══════════════════════════════════════════════════════════════════════════

import os as _os

# Folder containing input xlsx files (relative to this script, or absolute path)
INPUT_FOLDER  = _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), "input")

# Folder where the output xlsx will be saved
OUTPUT_FOLDER = _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), "output")

# Output file name
OUTPUT_FILE   = "hs_sector_mapped_combined.xlsx"

# Partners file settings (only relevant if you have Partner xlsx files)
PARTNER_COL   = "Country"          # column name for trading partner country
VALUE_COL     = None               # trade value column; None = auto-detect

# ───────────────────────────────────────────────────────────────────────────
# Everything below this line is the processing engine — no need to edit.
# ───────────────────────────────────────────────────────────────────────────

import sys
import re
import glob
import textwrap
import pandas as pd
import numpy as np

try:
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl is required. Install it with:  pip install openpyxl")
    sys.exit(1)


# ═══════════════════════════════════════════════════════════════════════════
# SECTION 1 — NIC SECTOR MAPPING
# ═══════════════════════════════════════════════════════════════════════════

ISIC_TO_SECTOR = {
    10: "Food and Agro",
    11: "Food and Agro",
    12: "Food and Agro",
    13: "Textiles",
    14: "Textiles",
    15: "Consumer Goods",
    16: "Construction Materials",
    17: "General Manufacturing",
    18: "General Manufacturing",
    19: "Chemicals",
    20: "Chemicals",
    21: "Chemicals",        # included in Chemicals total; also memo row
    22: "Chemicals",
    23: "Construction Materials",
    24: "Metals",
    25: "Metals",
    26: "Electronics",
    27: "Electrical Equipment",
    28: "Engineering Goods",
    29: "Transport Equipment",
    30: "Transport Equipment",
    31: "Construction Materials",
    32: "General Manufacturing",
}

SECTOR_ORDER = [
    "Food and Agro",
    "Consumer Goods",
    "Textiles",
    "Construction Materials",
    "Transport Equipment",
    "Engineering Goods",
    "Electrical Equipment",
    "Electronics",
    "Metals",
    "Chemicals",
    "General Manufacturing",
]

MEMO_SECTOR = "Pharmaceuticals (memo: NIC 21, within Chemicals)"
MEMO_NIC    = 21

ISIC_LABELS = {
    10: "Manufacture of food products",
    11: "Manufacture of beverages",
    12: "Manufacture of tobacco products",
    13: "Manufacture of textiles",
    14: "Manufacture of wearing apparel",
    15: "Manufacture of leather and related products",
    16: "Manufacture of wood and wood products (excl. furniture)",
    17: "Manufacture of paper and paper products",
    18: "Printing and reproduction of recorded media",
    19: "Manufacture of coke and refined petroleum products",
    20: "Manufacture of chemicals and chemical products",
    21: "Manufacture of pharmaceuticals, medicinal chemical & botanical products",
    22: "Manufacture of rubber and plastics products",
    23: "Manufacture of other non-metallic mineral products",
    24: "Manufacture of basic metals",
    25: "Manufacture of fabricated metal products (excl. machinery)",
    26: "Manufacture of computer, electronic and optical products",
    27: "Manufacture of electrical equipment",
    28: "Manufacture of machinery and equipment n.e.c.",
    29: "Manufacture of motor vehicles, trailers and semi-trailers",
    30: "Manufacture of other transport equipment",
    31: "Manufacture of furniture",
    32: "Other manufacturing",
}


# ═══════════════════════════════════════════════════════════════════════════
# SECTION 2 — FILE LOADERS
# ═══════════════════════════════════════════════════════════════════════════

def detect_year_columns(df: pd.DataFrame):
    """Return columns matching 'YYYY - YYYY' (e.g. '2024 - 2025')."""
    return [c for c in df.columns if re.match(r"^\d{4}\s*-\s*\d{4}$", str(c).strip())]


def load_tradestat_file(path: str) -> tuple:
    """
    Load one TradeStat commodity-wise export/import xlsx.

    Real TradeStat file format:
      Row 0 : "TradeStat->Eidb->Export->Commodity-wise"
      Row 1 : "Report Generated on: ... - Values in Rs Crore"
      Row 2 : S.No. | HSCode | Commodity | YYYY-YYYY | %Share | YYYY-YYYY | %Share | %Growth
      Row 3+ : data rows

    Returns (dataframe, [year_col_1, year_col_2]).
    Values are in Rs Crore as-is from the source file.
    """
    raw = pd.read_excel(path, header=None, dtype=str)

    # Find the row containing "HSCode" — works regardless of preamble row count
    header_row = None
    for i, row in raw.iterrows():
        if any("HSCode" in str(v) for v in row):
            header_row = i
            break
    if header_row is None:
        raise ValueError(f"Cannot find 'HSCode' header in {_os.path.basename(path)}")

    df = pd.read_excel(path, skiprows=header_row, header=0, dtype=str)
    # Normalise names; pandas auto-suffixes duplicate %Share columns (%Share.1 etc.)
    df.columns = [str(c).strip() for c in df.columns]

    # Keep only rows where HSCode is exactly 6 digits — drops totals, blanks, notes
    df = df[df["HSCode"].str.match(r"^\d{6}$", na=False)].copy()
    df.reset_index(drop=True, inplace=True)

    year_cols = detect_year_columns(df)
    if not year_cols:
        raise ValueError(
            f"No year columns (YYYY - YYYY) found in {_os.path.basename(path)}.\n"
            f"  Columns found: {list(df.columns)}"
        )

    return df, year_cols


def load_partners_file(path: str,
                       partner_col: str = "Country",
                       value_col: str   = None,
                       year_label: str  = None) -> tuple:
    """
    Load a bilateral trading-partners xlsx.
    Returns (dataframe, [year_label], partner_col, value_col).
    """
    raw = pd.read_excel(path, header=None, dtype=str)
    header_row = None
    for i, row in raw.iterrows():
        if any("HSCode" in str(v) for v in row):
            header_row = i
            break
    if header_row is None:
        header_row = 0

    df = pd.read_excel(path, skiprows=header_row, header=0, dtype=str)
    df.columns = [str(c).strip() for c in df.columns]

    if "HSCode" not in df.columns:
        raise ValueError(
            f"'HSCode' column not found in {_os.path.basename(path)}. "
            f"Columns: {list(df.columns)}"
        )

    df = df[df["HSCode"].str.match(r"^\d{6}$", na=False)].copy()
    df.reset_index(drop=True, inplace=True)

    # Auto-detect value column if not specified
    if value_col is None:
        skip = {"HSCode", partner_col}
        for col in df.columns:
            if col in skip:
                continue
            try:
                numeric = pd.to_numeric(df[col], errors="coerce")
                if numeric.notna().sum() > 0:
                    value_col = col
                    break
            except Exception:
                continue
        if value_col is None:
            raise ValueError(
                f"Could not auto-detect a numeric value column in {_os.path.basename(path)}. "
                f"Please set VALUE_COL in the CONFIG section."
            )
        print(f"    [auto-detected value column: '{value_col}']")

    if year_label is None:
        year_label = _os.path.splitext(_os.path.basename(path))[0]

    return df, [year_label], partner_col, value_col


# ═══════════════════════════════════════════════════════════════════════════
# SECTION 3 — HS → ISIC REV 4 ASSIGNMENT
# Direct translation of Stata .do file; later rules overwrite earlier ones.
# ═══════════════════════════════════════════════════════════════════════════

def assign_isic4(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df["_cc"]      = df["HSCode"].astype(int)
    df["HS2digit"] = df["HSCode"].str[:2].astype(int)
    td = df["HS2digit"]
    cc = df["_cc"]

    # ── Drop flags (non-manufacturing) ──────────────────────────────────
    keep  = cc > 20000
    keep &= ~((cc >= 60110)  & (cc <= 60499))
    keep &= ~((cc >= 130110) & (cc <= 130239))
    keep &= ~((cc >= 270111) & (cc <= 270300))
    keep &= ~cc.isin([270500, 270900, 271410, 271490, 411520])
    keep &= ~((cc >= 391510) & (cc <= 391590))
    keep &= ~((cc >= 711230) & (cc <= 711299))
    keep &= ~((cc >= 400110) & (cc <= 400130))
    keep &= ~((cc >= 410320) & (cc <= 410390))
    keep &= ~cc.isin([710110, 710121])
    keep &= ~td.isin([97, 99])
    keep &= (cc != 271600)
    df["_keep"] = keep

    isic = pd.Series(pd.NA, index=df.index, dtype="Float64")

    # 10 — Food products
    isic[td.isin([2,3,4,5,7,8,9,10,11,12,14,15,16,17,18,19,20,21,23])] = 10
    isic[(cc >= 410110) & (cc <= 410310)] = 10
    isic[cc.isin([350110,350190,350211,350219,350510])] = 10

    # 11 — Beverages
    isic[td == 22] = 11

    # 12 — Tobacco
    isic[td == 24] = 12

    # 13 — Textiles
    isic[td.isin([50,51,52,53,54,55,56,57,58,59,60,63,94])] = 13
    isic[cc.isin([701940,701951,701952,701959])] = 13
    isic[cc == 880400] = 13

    # 14 — Wearing apparel
    isic[td.isin([61,62,65,43])] = 14
    isic[cc.isin([420310,420329,420330,420340])] = 14

    # 15 — Leather & related
    isic[(cc >= 410410) & (cc <= 411510)] = 15
    isic[(cc >= 420100) & (cc <= 420299)] = 15
    isic[cc.isin([420400,420500,911390,960500])] = 15
    isic[td == 64] = 15

    # 16 — Wood & wood products
    isic[td.isin([44,45,46])] = 16

    # 17 — Paper & paper products
    isic[td == 47] = 17
    isic[(cc >= 480100) & (cc <= 482390)] = 17
    isic[cc == 590500] = 17

    # 18 — Printing & recorded media
    isic[td == 49] = 18
    isic[(cc >= 482010) & (cc <= 482090)] = 18
    isic[cc.isin([844250,852351,852359,852380])] = 18

    # 19 — Coke & refined petroleum
    isic[(cc >= 271000) & (cc <= 271390)] = 19
    isic[cc.isin([270400,270600])] = 19
    isic[cc.isin([284410,284420,284430,284440])] = 19

    # 20 — Chemicals
    isic[td.isin([26,29,32,31,33,34,37])] = 20
    isic[(td == 35) & isic.isna()] = 20
    isic[(cc >= 270710) & (cc <= 270820)] = 20
    isic[(cc >= 280110) & (cc <= 284390)] = 20
    isic[(cc >= 284510) & (cc <= 285300)] = 20
    isic[(cc >= 380210) & (cc <= 382590)] = 20
    isic[cc.isin([440200,300670,710410,710420])] = 20
    isic[(cc >= 390110) & (cc <= 391400)] = 20
    isic[(cc >= 400211) & (cc <= 400299)] = 20
    isic[(cc >= 360100) & (cc <= 360490)] = 20
    isic[(cc >= 852321) & (cc <= 852340)] = 20
    isic[(cc >= 540210) & (cc <= 540259)] = 20
    isic[(cc >= 540310) & (cc <= 540339)] = 20
    isic[(cc >= 540410) & (cc <= 540500)] = 20
    isic[(cc >= 550110) & (cc <= 550490)] = 20
    isic[cc.isin([151800,152000])] = 20

    # 21 — Pharmaceuticals
    isic[cc.isin([291821,291822,291823,300692])] = 21
    isic[cc.isin([292241,292242])] = 21
    isic[(cc >= 292310) & (cc <= 292410)] = 21
    isic[cc.isin([292422,292429,293229,293311,293319,293321,293430])] = 21
    isic[(cc >= 293351) & (cc <= 293369)] = 21
    isic[(cc >= 293500) & (cc <= 294190)] = 21
    isic[(cc >= 300120) & (cc <= 300660)] = 21

    # 22 — Rubber & plastics
    isic[(cc >= 391610) & (cc <= 392690)] = 22
    isic[(cc >= 400300) & (cc <= 401700)] = 22
    isic[cc.isin([590610,300691,590691,590699,853670])] = 22
    isic[cc.isin([650610,650691,854720,940592])] = 22

    # 23 — Non-metallic mineral products
    isic[td.isin([25,68,69])] = 23
    isic[cc.isin([271500,281810,380110,380120,380130,380190,
                  381600,382450,854610,854620,854710])] = 23
    isic[(td == 70) & isic.isna()] = 23

    # 24 — Basic metals
    isic[td.isin([72,75,78,79,80,81])] = 24
    isic[(cc >= 730110) & (cc <= 730799)] = 24
    isic[(cc >= 740110) & (cc <= 741220)] = 24
    isic[(cc >= 760110) & (cc <= 760900)] = 24
    isic[(cc >= 710610) & (cc <= 711100)] = 24
    isic[cc == 281820] = 24

    # 25 — Fabricated metal products (overwrites copper 74 range intentionally)
    isic[(cc >= 730810) & (cc <= 731450)] = 25
    isic[(cc >= 731520) & (cc <= 732090)] = 25
    isic[(cc >= 732211) & (cc <= 732290)] = 25
    isic[(cc >= 732310) & (cc <= 732690)] = 25
    isic[(td == 74) & (cc != 741700)] = 25
    isic[td.isin([82,83])] = 25
    isic[cc.isin([750810,750890,780600,790700,800700,
                  840110,840410,840420,840490,848710,940600])] = 25
    isic[(cc >= 761010) & (cc <= 761699)] = 25
    isic[(cc >= 840140) & (cc <= 840290)] = 25

    # 26 — Computer, electronic & optical
    isic[cc.isin([841920,844312,852352,940210,940290])] = 26
    isic[cc.isin([844331,844332,844339,844399])] = 26
    isic[(cc >= 846900) & (cc <= 847350)] = 26
    isic[(cc >= 851711) & (cc <= 852290)] = 26
    isic[(cc >= 852510) & (cc <= 852990)] = 26
    isic[(cc >= 853210) & (cc <= 853400)] = 26
    isic[(cc >= 854011) & (cc <= 854290)] = 26
    isic[(td == 90) & (cc != 902300)] = 26
    isic[(td == 91) & (cc != 911390)] = 26

    # 27 — Electrical equipment
    isic[cc.isin([630110,732290,741700,840310,840390,841451,841460,
                  841911,841919,842211,842219,845121,854690,940599,900662])] = 27
    isic[(cc >= 732111) & (cc <= 732190)] = 27
    isic[(cc >= 841810) & (cc <= 841840)] = 27
    isic[(cc >= 845011) & (cc <= 845019)] = 27
    isic[(cc >= 940510) & (cc <= 940560)] = 27
    isic[(cc >= 850110) & (cc <= 850790)] = 27
    isic[(cc >= 850811) & (cc <= 851390)] = 27
    isic[(cc >= 851610) & (cc <= 851690)] = 27
    isic[(cc >= 853010) & (cc <= 853190)] = 27
    isic[(cc >= 853510) & (cc <= 853990)] = 27  # overwrites 853670 from isic22
    isic[(cc >= 854311) & (cc <= 854590)] = 27
    isic[(cc >= 854790) & (cc <= 854890)] = 27

    # 28 — Machinery & equipment n.e.c.
    isic[cc.isin([731511,731512,731519,840130,840120,840721,840729,840790,
                  840810,871620,844391,840890,841181,841182,841199,848790,
                  841459,844311,870110,870130,870190,854310])] = 28
    isic[(cc >= 840510) & (cc <= 840690)] = 28
    isic[(cc >= 841011) & (cc <= 841090)] = 28
    isic[(cc >= 841480) & (cc <= 841790)] = 28
    isic[(cc >= 841221) & (cc <= 841440)] = 28
    isic[(cc >= 841850) & (cc <= 841899)] = 28
    isic[(cc >= 841931) & (cc <= 842199)] = 28
    isic[(cc >= 842220) & (cc <= 844240)] = 28
    isic[(cc >= 844313) & (cc <= 844319)] = 28
    isic[(cc >= 844400) & (cc <= 844900)] = 28
    isic[(cc >= 845020) & (cc <= 845110)] = 28
    isic[(cc >= 845129) & (cc <= 846890)] = 28
    isic[(cc >= 847410) & (cc <= 848690)] = 28
    isic[(cc >= 850810) & (cc <= 850890)] = 28  # overwrites isic27 range
    isic[(cc >= 851410) & (cc <= 851590)] = 28
    isic[(cc >= 870911) & (cc <= 871000)] = 28
    isic[td == 93] = 28

    # 29 — Motor vehicles
    isic[cc.isin([840820,840991,840999,870120,871610,871690])] = 29
    isic[(cc >= 840731) & (cc <= 840734)] = 29
    isic[(cc >= 870210) & (cc <= 870899)] = 29
    isic[(cc >= 871631) & (cc <= 871640)] = 29

    # 30 — Other transport equipment
    isic[cc.isin([840710,871680,840910,841191,841210])] = 30
    isic[td == 86] = 30
    isic[td == 89] = 30
    isic[(td == 88) & (cc != 880400)] = 30
    isic[(cc >= 841111) & (cc <= 841122)] = 30
    isic[(cc >= 871110) & (cc <= 871499)] = 30

    # 31 — Furniture
    isic[(cc >= 940110) & (cc <= 940190)] = 31
    isic[(cc >= 940310) & (cc <= 940429)] = 31

    # 32 — Other manufacturing (unconditional; overwrites earlier assignments for td 96)
    isic[td.isin([66,67,92,95,96])] = 32
    isic[cc.isin([340600,360500,360610,360690,420600,420610,420690,
                  590410,590491,590492,420321,871500,902300,
                  711711,711719,711790])] = 32
    isic[(cc >= 710122) & (cc <= 710399)] = 32
    isic[(cc >= 710490) & (cc <= 710590)] = 32
    isic[(cc >= 711311) & (cc <= 711890)] = 32

    df["ISIC4"] = isic
    return df


# ═══════════════════════════════════════════════════════════════════════════
# SECTION 4 — PROCESS FILES → long-format rows
# ═══════════════════════════════════════════════════════════════════════════

def _tag_sector(df: pd.DataFrame) -> pd.DataFrame:
    df["NIC_Sector"] = df["ISIC4"].map(
        lambda x: ISIC_TO_SECTOR.get(int(x), "Non-Manufacturing") if pd.notna(x) else "Non-Manufacturing"
    )
    df["ISIC4_Label"] = df["ISIC4"].map(
        lambda x: ISIC_LABELS.get(int(x), "Unmapped") if pd.notna(x) else "Not Mapped"
    )
    return df


def process_file(path: str) -> pd.DataFrame:
    df, year_cols = load_tradestat_file(path)
    df = assign_isic4(df)
    df = _tag_sector(df)

    mfg = df[df["_keep"] & df["ISIC4"].notna()].copy()

    id_cols = ["HSCode", "Commodity", "ISIC4", "ISIC4_Label", "NIC_Sector"]
    rows = []
    for ycol in year_cols:
        tmp = mfg[id_cols + [ycol]].copy()
        tmp["Year"]        = ycol.strip()
        tmp["Value_INRCr"] = pd.to_numeric(mfg[ycol], errors="coerce").fillna(0)
        rows.append(tmp[id_cols + ["Year", "Value_INRCr"]])

    return pd.concat(rows, ignore_index=True)


def process_partners_file(path: str,
                           partner_col: str = "Country",
                           value_col: str   = None,
                           year_label: str  = None) -> pd.DataFrame:
    df, year_labels, pcol, vcol = load_partners_file(
        path, partner_col=partner_col, value_col=value_col, year_label=year_label
    )
    df = assign_isic4(df)
    df = _tag_sector(df)

    mfg = df[df["_keep"] & df["ISIC4"].notna()].copy()

    id_cols = ["HSCode", "ISIC4", "ISIC4_Label", "NIC_Sector", pcol]
    if "Commodity" in mfg.columns:
        id_cols = ["HSCode", "Commodity", "ISIC4", "ISIC4_Label", "NIC_Sector", pcol]

    mfg = mfg.copy()
    mfg["Year"]        = year_labels[0]
    mfg["Value_INRCr"] = pd.to_numeric(mfg[vcol], errors="coerce").fillna(0)

    return mfg[id_cols + ["Year", "Value_INRCr"]].copy()


# ═══════════════════════════════════════════════════════════════════════════
# SECTION 5 — BUILD SUMMARIES
# ═══════════════════════════════════════════════════════════════════════════

def build_sector_summary(long: pd.DataFrame) -> tuple:
    """Sector × Year pivot. Pharmaceuticals appended as memo (not in total)."""
    years = sorted(long["Year"].unique())

    grp = (
        long.groupby(["NIC_Sector", "Year"])["Value_INRCr"]
        .sum().reset_index()
    )
    pivot = grp.pivot(index="NIC_Sector", columns="Year", values="Value_INRCr").fillna(0)

    for s in SECTOR_ORDER:
        if s not in pivot.index:
            pivot.loc[s] = 0
    pivot = pivot.reindex(SECTOR_ORDER)

    total = pivot.sum()
    total.name = "TOTAL MANUFACTURING"

    pharma = (
        long[long["ISIC4"] == MEMO_NIC]
        .groupby("Year")["Value_INRCr"]
        .sum()
        .reindex(years, fill_value=0)
    )
    pharma.name = MEMO_SECTOR

    out = pd.concat([pivot, total.to_frame().T, pharma.to_frame().T])
    out = out[years]
    out.index.name = "NIC Sector (IME)"
    out = out.reset_index()

    for y in years:
        out[y] = out[y].round(2)

    # Add Share (%) column for every year, and YoY Growth (%) for every consecutive pair.
    # Share = sector value / TOTAL MANUFACTURING for that year * 100.
    # YoY   = (curr - prev) / prev * 100  — standard period-over-period growth formula.
    for y in years:
        total_y = out.loc[out["NIC Sector (IME)"] == "TOTAL MANUFACTURING", y].values[0]
        if total_y != 0:
            out[f"Share {y} (%)"] = (out[y] / total_y * 100).round(2)
            out.loc[out["NIC Sector (IME)"] == "TOTAL MANUFACTURING", f"Share {y} (%)"] = 100.0
        else:
            out[f"Share {y} (%)"] = np.nan

    for i in range(1, len(years)):
        prev_y, curr_y = years[i - 1], years[i]
        col_label = f"YoY Growth {prev_y}→{curr_y} (%)"
        out[col_label] = (
            ((out[curr_y] - out[prev_y]) / out[prev_y].replace(0, np.nan)) * 100
        ).round(2)

    return out, years


def build_isic_summary(long: pd.DataFrame, years: list) -> pd.DataFrame:
    grp = (
        long.groupby(["ISIC4", "ISIC4_Label", "NIC_Sector", "Year"])["Value_INRCr"]
        .sum().reset_index()
    )
    pivot = grp.pivot_table(
        index=["ISIC4", "ISIC4_Label", "NIC_Sector"],
        columns="Year", values="Value_INRCr", fill_value=0
    ).reset_index()
    pivot.columns.name = None
    pivot["ISIC4"] = pivot["ISIC4"].astype(int)
    pivot = pivot.sort_values("ISIC4").reset_index(drop=True)
    pivot = pivot.rename(columns={
        "ISIC4":      "ISIC4 (2-digit)",
        "ISIC4_Label": "ISIC4 Description",
        "NIC_Sector":  "NIC Sector (IME)",
    })
    for y in years:
        if y in pivot.columns:
            pivot[y] = pivot[y].round(2)

    pivot["Note"] = ""
    pivot.loc[pivot["ISIC4 (2-digit)"] == 21, "Note"] = (
        "NIC 21 (Pharma) counted inside Chemicals total "
        "AND extracted separately as memo row in Sector Summary"
    )
    return pivot


def build_hs_detail(long: pd.DataFrame, years: list) -> pd.DataFrame:
    index_cols = ["HSCode", "ISIC4", "ISIC4_Label", "NIC_Sector"]
    if "Commodity" in long.columns:
        index_cols = ["HSCode", "Commodity", "ISIC4", "ISIC4_Label", "NIC_Sector"]

    pivot = long.pivot_table(
        index=index_cols,
        columns="Year", values="Value_INRCr", fill_value=0
    ).reset_index()
    pivot.columns.name = None
    pivot["ISIC4"] = pivot["ISIC4"].astype(int)
    pivot = pivot.rename(columns={
        "ISIC4":       "ISIC4 (2-digit)",
        "ISIC4_Label": "ISIC4 Description",
        "NIC_Sector":  "NIC Sector (IME)",
    })
    pivot = pivot.sort_values(["NIC Sector (IME)", "ISIC4 (2-digit)", "HSCode"]).reset_index(drop=True)
    for y in years:
        if y in pivot.columns:
            pivot[y] = pivot[y].round(4)
    return pivot


def build_partners_summary(long: pd.DataFrame, partner_col: str = "Country") -> tuple:
    years = sorted(long["Year"].unique())
    grp = (
        long.groupby(["NIC_Sector", partner_col, "Year"])["Value_INRCr"]
        .sum().reset_index()
    )
    out_frames = []
    for year in years:
        yd = grp[grp["Year"] == year].pivot(
            index="NIC_Sector", columns=partner_col, values="Value_INRCr"
        ).fillna(0)
        yd.columns = [f"{c} ({year})" for c in yd.columns]
        out_frames.append(yd)

    combined = pd.concat(out_frames, axis=1).fillna(0)
    for s in SECTOR_ORDER:
        if s not in combined.index:
            combined.loc[s] = 0
    combined = combined.reindex(SECTOR_ORDER)
    combined.index.name = "NIC Sector (IME)"
    combined = combined.reset_index()
    return combined, years


def build_year_on_year_trends(long: pd.DataFrame, years: list) -> pd.DataFrame:
    """
    NEW: Year-on-year growth table for every sector across all detected years.
    Useful for spotting trends across the full 2018–2025 span.
    """
    grp = (
        long.groupby(["NIC_Sector", "Year"])["Value_INRCr"]
        .sum().reset_index()
    )
    pivot = grp.pivot(index="NIC_Sector", columns="Year", values="Value_INRCr").fillna(0)
    for s in SECTOR_ORDER:
        if s not in pivot.index:
            pivot.loc[s] = 0
    pivot = pivot.reindex(SECTOR_ORDER)
    total = pivot.sum()
    total.name = "TOTAL MANUFACTURING"
    pivot = pd.concat([pivot, total.to_frame().T])
    pivot = pivot[years]

    # Compute YoY growth for each consecutive pair
    growth_cols = {}
    for i in range(1, len(years)):
        prev, curr = years[i-1], years[i]
        label = f"Growth {prev}→{curr} (%)"
        growth_cols[label] = ((pivot[curr] - pivot[prev]) / pivot[prev].replace(0, np.nan) * 100).round(2)

    result = pivot.copy()
    for lbl, col in growth_cols.items():
        result[lbl] = col

    result.index.name = "NIC Sector (IME)"
    result = result.reset_index()
    for y in years:
        result[y] = result[y].round(2)
    return result


# ═══════════════════════════════════════════════════════════════════════════
# SECTION 6 — EXCEL WRITER WITH FORMATTING
# ═══════════════════════════════════════════════════════════════════════════

NAVY    = "1F3864"
TEAL    = "17375E"
AMBER   = "FFC000"
LIGHT   = "D6E4F0"
MEMO_BG = "FFF2CC"
TOTAL_BG= "BDD7EE"
GROWTH_POS = "E2EFDA"   # light green  — positive growth
GROWTH_NEG = "FFDDC1"   # light orange — negative growth
WHITE   = "FFFFFF"


def _header_style(ws, fill_hex=NAVY):
    fill  = PatternFill("solid", fgColor=fill_hex)
    font  = Font(bold=True, color=WHITE, name="Arial", size=10)
    align = Alignment(wrap_text=True, vertical="center", horizontal="center")
    for cell in ws[1]:
        cell.fill      = fill
        cell.font      = font
        cell.alignment = align
    ws.row_dimensions[1].height = 32


def _autowidth(ws, max_width=55):
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        width = max(len(str(c.value or "")) for c in col_cells)
        ws.column_dimensions[col_letter].width = min(width + 3, max_width)


def _freeze(ws, cell="B2"):
    ws.freeze_panes = cell


def write_output(df_sector, df_isic, df_detail, df_trends, years, output_path,
                 df_partners=None, partners_sheet_name="Partner Summary"):
    """
    Write all summary dataframes into one Excel workbook.

    Sheets
    ──────
    1. Sector Summary        — 11 IME sectors + total + Pharma memo
    2. Year-on-Year Trends   — all years side-by-side with growth rates  ← NEW
    3. ISIC Summary          — ISIC 2-digit level
    4. HS Detail             — every HS code
    5. Partner Summary       — optional, if partners files supplied
    6. Mapping Reference     — NIC → sector crosswalk
    """
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:

        # ── Sheet 1: Sector Summary ────────────────────────────────────────
        df_sector.to_excel(writer, sheet_name="Sector Summary", index=False)
        ws = writer.sheets["Sector Summary"]
        _header_style(ws)
        _autowidth(ws)
        _freeze(ws, "B2")
        ss_headers = [cell.value for cell in ws[1]]
        ss_growth_cols = [i for i, h in enumerate(ss_headers) if h and "YoY Growth" in str(h)]
        for row in ws.iter_rows(min_row=2):
            val = str(row[0].value or "")
            if val == "TOTAL MANUFACTURING":
                for cell in row:
                    cell.fill = PatternFill("solid", fgColor=TOTAL_BG)
                    cell.font = Font(bold=True, name="Arial", size=10)
            elif MEMO_SECTOR in val:
                for cell in row:
                    cell.fill = PatternFill("solid", fgColor=MEMO_BG)
                    cell.font = Font(italic=True, name="Arial", size=10)
                for cell in row:
                    cell.border = Border(top=Side(style="medium", color=AMBER))
            else:
                for cell in row:
                    cell.font = Font(name="Arial", size=10)
            # Colour-code YoY Growth columns green/orange
            for ci in ss_growth_cols:
                cell = row[ci]
                try:
                    v = float(cell.value) if cell.value is not None else None
                    if v is not None:
                        cell.fill = PatternFill("solid", fgColor=GROWTH_POS if v >= 0 else GROWTH_NEG)
                except (TypeError, ValueError):
                    pass

        # ── Sheet 2: Year-on-Year Trends (NEW) ────────────────────────────
        df_trends.to_excel(writer, sheet_name="Year-on-Year Trends", index=False)
        ws_t = writer.sheets["Year-on-Year Trends"]
        _header_style(ws_t, fill_hex="2E4057")
        _autowidth(ws_t)
        _freeze(ws_t, "B2")
        # Colour growth cells green / red
        headers = [cell.value for cell in ws_t[1]]
        growth_col_indices = [i for i, h in enumerate(headers) if h and "Growth" in str(h)]
        for row in ws_t.iter_rows(min_row=2):
            val = str(row[0].value or "")
            base_font = Font(bold=True, name="Arial", size=10) if val in ("TOTAL MANUFACTURING",) \
                        else Font(name="Arial", size=10)
            for cell in row:
                cell.font = base_font
            for ci in growth_col_indices:
                cell = row[ci]
                try:
                    v = float(cell.value) if cell.value is not None else None
                    if v is not None:
                        color = GROWTH_POS if v >= 0 else GROWTH_NEG
                        cell.fill = PatternFill("solid", fgColor=color)
                except (TypeError, ValueError):
                    pass

        # ── Sheet 3: ISIC Summary ──────────────────────────────────────────
        df_isic.to_excel(writer, sheet_name="ISIC Summary", index=False)
        ws2 = writer.sheets["ISIC Summary"]
        _header_style(ws2, fill_hex=TEAL)
        _autowidth(ws2)
        _freeze(ws2, "C2")
        for row in ws2.iter_rows(min_row=2):
            if row[0].value == 21:
                for cell in row:
                    cell.fill = PatternFill("solid", fgColor=MEMO_BG)
                    cell.font = Font(italic=True, name="Arial", size=10)
            else:
                for cell in row:
                    cell.font = Font(name="Arial", size=10)

        # ── Sheet 4: HS Detail ─────────────────────────────────────────────
        df_detail.to_excel(writer, sheet_name="HS Detail", index=False)
        ws3 = writer.sheets["HS Detail"]
        _header_style(ws3, fill_hex=TEAL)
        _autowidth(ws3)
        _freeze(ws3, "C2")
        for row in ws3.iter_rows(min_row=2):
            for cell in row:
                cell.font = Font(name="Arial", size=10)

        # ── Sheet 5: Partners Summary (optional) ──────────────────────────
        if df_partners is not None:
            df_partners.to_excel(writer, sheet_name=partners_sheet_name, index=False)
            ws_p = writer.sheets[partners_sheet_name]
            _header_style(ws_p, fill_hex=NAVY)
            _autowidth(ws_p)
            _freeze(ws_p, "B2")
            for row in ws_p.iter_rows(min_row=2):
                for cell in row:
                    cell.font = Font(name="Arial", size=10)

        # ── Sheet 6: Mapping Reference ─────────────────────────────────────
        ref_data = []
        for isic_code in sorted(ISIC_TO_SECTOR.keys()):
            note = ""
            if isic_code == 21:
                note = "Also extracted as Pharmaceuticals memo row (NOT added to total)"
            ref_data.append({
                "ISIC4 (2-digit)":   isic_code,
                "ISIC4 Description": ISIC_LABELS[isic_code],
                "NIC Sector (IME)":  ISIC_TO_SECTOR[isic_code],
                "Note":              note,
            })
        df_ref = pd.DataFrame(ref_data)
        df_ref.to_excel(writer, sheet_name="Mapping Reference", index=False)
        ws4 = writer.sheets["Mapping Reference"]
        _header_style(ws4, fill_hex=NAVY)
        _autowidth(ws4)
        for row in ws4.iter_rows(min_row=2):
            if row[0].value == 21:
                for cell in row:
                    cell.fill = PatternFill("solid", fgColor=MEMO_BG)
                    cell.font = Font(italic=True, name="Arial", size=10)
            else:
                for cell in row:
                    cell.font = Font(name="Arial", size=10)


# ═══════════════════════════════════════════════════════════════════════════
# SECTION 7 — MAIN PIPELINE
# ═══════════════════════════════════════════════════════════════════════════

def run(input_files: list,
        output_path: str        = "hs_sector_mapped_combined.xlsx",
        partners_files: list    = None,
        partner_col: str        = "Country",
        value_col: str          = None):
    """
    Full HS → ISIC → Sector pipeline.

    Parameters
    ----------
    input_files    : list of Tradestat trade-volume xlsx paths (multiple years OK)
    output_path    : destination for the combined output workbook
    partners_files : list of bilateral trading-partners xlsx paths (optional)
    partner_col    : column name for partner country (default: 'Country')
    value_col      : trade value column name (auto-detected if None)
    """
    all_long = []

    # ── Trade volume files ─────────────────────────────────────────────────
    if input_files:
        print(f"\nProcessing {len(input_files)} trade-volume file(s)...")
        for fp in sorted(input_files):
            print(f"  → {_os.path.basename(fp)}")
            try:
                rows = process_file(fp)
                all_long.append(rows)
                print(f"     {len(rows):,} HS-year rows loaded")
            except Exception as e:
                print(f"     ERROR: {e} — skipping file")

    # ── Partners files ─────────────────────────────────────────────────────
    partners_long = []
    if partners_files:
        print(f"\nProcessing {len(partners_files)} trading-partners file(s)...")
        for fp in sorted(partners_files):
            print(f"  → {_os.path.basename(fp)}")
            try:
                rows = process_partners_file(fp, partner_col=partner_col, value_col=value_col)
                partners_long.append(rows)
                trade_cols = ["HSCode"] + \
                             (["Commodity"] if "Commodity" in rows.columns else []) + \
                             ["ISIC4", "ISIC4_Label", "NIC_Sector", "Year", "Value_INRCr"]
                all_long.append(rows[trade_cols])
                print(f"     {len(rows):,} HS-partner rows loaded")
            except Exception as e:
                print(f"     ERROR: {e} — skipping file")

    if not all_long:
        print("\nNo data loaded — nothing to process.")
        print("  Trade volume : name files containing 'Trade'  e.g. TradeStat_2022-23.xlsx")
        print("  Partners     : name files containing 'Partner' e.g. Partners_2022-23.xlsx")
        return

    long = pd.concat(all_long, ignore_index=True)

    # ── De-duplicate overlapping year columns ─────────────────────────────────────
    # TradeStat files each contain TWO year columns: the current year AND the prior
    # year for comparison. When you load multiple consecutive files (e.g. 2024-25
    # and 2025-26), the shared year (2024-25) appears in both files with identical
    # values. Summing would double-count it. Instead, keep exactly one copy per
    # (HSCode, Year) pair — the first occurrence is fine since values are identical
    # across files for the same year.
    id_cols_dedup = [c for c in ["HSCode", "Commodity", "ISIC4", "ISIC4_Label", "NIC_Sector", "Year"]
                     if c in long.columns]
    long = long.drop_duplicates(subset=id_cols_dedup, keep="first").copy()
    long = (
        long.groupby(id_cols_dedup, dropna=False)["Value_INRCr"]
        .sum().reset_index()
    )

    years = sorted(long["Year"].unique())
    print(f"\nYears detected   : {years}")
    print(f"Total HS-yr rows : {len(long):,}")

    print("\nBuilding summaries...")
    df_sector, years = build_sector_summary(long)
    df_isic           = build_isic_summary(long, years)
    df_detail         = build_hs_detail(long, years)
    df_trends         = build_year_on_year_trends(long, years)   # NEW

    df_partners_out = None
    if partners_long:
        pl = pd.concat(partners_long, ignore_index=True)
        df_partners_out, _ = build_partners_summary(pl, partner_col=partner_col)

    print(f"\nWriting output → {output_path}")
    _os.makedirs(_os.path.dirname(output_path), exist_ok=True)
    write_output(df_sector, df_isic, df_detail, df_trends, years, output_path,
                 df_partners=df_partners_out)

    print("\n✓ Done!\n")
    print("=== SECTOR SUMMARY (₹ Crore) ===")
    print(df_sector.to_string(index=False))
    print()
    print("Sheets written:")
    print("  1. Sector Summary       — 11 IME sectors + total + Pharmaceuticals memo row")
    print("  2. Year-on-Year Trends  — all years side-by-side with colour-coded growth rates")
    print("  3. ISIC Summary         — ISIC4 2-digit level (NIC 21 highlighted)")
    print("  4. HS Detail            — every HS code with ISIC4 and sector, all years")
    if df_partners_out is not None:
        print("  5. Partner Summary      — sector × partner breakdown")
        print("  6. Mapping Reference    — full NIC → sector crosswalk")
    else:
        print("  5. Mapping Reference    — full NIC → sector crosswalk")

    return df_sector


# ═══════════════════════════════════════════════════════════════════════════
# ENTRY POINT — runs when you execute: python hs_sector_mapper_vsc.py
# ═══════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    print("=" * 65)
    print("  IME Manufacturing Trade Mapper — VSCode / Local Edition")
    print("=" * 65)

    # ── Create input/output folders if they don't exist ───────────────────
    for folder in (INPUT_FOLDER, OUTPUT_FOLDER):
        if not _os.path.exists(folder):
            _os.makedirs(folder)
            print(f"  Created folder: {folder}")

    # ── Discover input files ───────────────────────────────────────────────
    # Matches real TradeStat filenames like:
    #   TradeStat-Eidb-Export-Commodity-wise_2024-25.xlsx
    #   TradeStat-Eidb-Import-Commodity-wise_2023-24.xlsx
    # Also catches any xlsx with "trade" or "tradestat" in the name.
    all_xlsx = glob.glob(_os.path.join(INPUT_FOLDER, "*.xlsx"))

    def _is_trade_file(name: str) -> bool:
        n = name.lower()
        return "tradestat" in n or n.startswith("trade")

    def _is_partner_file(name: str) -> bool:
        return "partner" in _os.path.basename(name).lower()

    input_files    = [f for f in all_xlsx if _is_trade_file(_os.path.basename(f))]
    partners_files = [f for f in all_xlsx if _is_partner_file(f)]

    # Fallback: if still nothing matched, treat ALL xlsx as trade files
    if not input_files and not partners_files and all_xlsx:
        print("\n  NOTE: No files matched the expected naming pattern.")
        print("  Treating all xlsx files in the input folder as Tradestat files.")
        input_files = all_xlsx

    if not input_files and not partners_files:
        print(f"\n  No xlsx files found in: {INPUT_FOLDER}")
        print(textwrap.dedent("""
  ── HOW TO ADD FILES ─────────────────────────────────────────────
  1. Place your xlsx files in the "input" folder next to this script.
  2. Files must start with "TradeStat" (any suffix is fine):
       TradeStat-Eidb-Export-Commodity-wise_2018-19.xlsx
       TradeStat-Eidb-Export-Commodity-wise_2019-20.xlsx
       ...up to 2024-25 or 2025-26
  3. Partner files (optional): include "Partner" in the filename.
  4. Run the script again.
  ─────────────────────────────────────────────────────────────────
        """))
        sys.exit(0)

    print(f"\n  Input folder  : {INPUT_FOLDER}")
    print(f"  Trade files   : {len(input_files)}")
    print(f"  Partner files : {len(partners_files)}")

    output_path = _os.path.join(OUTPUT_FOLDER, OUTPUT_FILE)

    run(
        input_files    = input_files,
        output_path    = output_path,
        partners_files = partners_files or None,
        partner_col    = PARTNER_COL,
        value_col      = VALUE_COL,
    )

    print(f"\n  Output saved to: {output_path}")

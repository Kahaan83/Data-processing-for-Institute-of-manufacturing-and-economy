"""
filter_top_companies.py
-----------------------
Filters top-N companies per selected NIC codes, ranked by sales for a chosen year
or across a range / list of years.  Each year gets its own ranking sheet; a cross-
year summary sheet is appended at the end.

Input : combined_company_sales.xlsx  (or a CSV pre-exported from it)
Output: <Sector>_FY<start>-FY<end>_top<N>.xlsx

Usage
-----
  python filter_top_companies.py                               # interactive prompts
  python filter_top_companies.py --file data.xlsx              # custom file path
  python filter_top_companies.py --csv                         # use/create CSV cache (fast)

  # Single year
  python filter_top_companies.py --sector "Steel" --nic "24" --years 2024 --top 10

  # Explicit list
  python filter_top_companies.py --sector "Steel" --nic "24" --years 2020,2022,2024 --top 10

  # Range  (start-end, inclusive)
  python filter_top_companies.py --sector "Steel" --nic "24" --years 2015-2024 --top 10
"""

import argparse
import os
import sys
import time

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ──────────────────────────────────────────────────────────────────────────────
AVAILABLE_YEARS = list(range(2014, 2027))   # 2014 – 2026

SORT_CHOICES = {
    "1": ("total",    "Total Sales (Products + Services)"),
    "2": ("products", "Products Sales only"),
    "3": ("services", "Services Sales only"),
}

# shared style constants (module-level so helpers can use them)
NAVY   = PatternFill("solid", fgColor="1F3864")
BLUE   = PatternFill("solid", fgColor="2E75B6")
GREEN  = PatternFill("solid", fgColor="70AD47")
ORANGE = PatternFill("solid", fgColor="C55A11")
GOLD   = PatternFill("solid", fgColor="FFD700")
SILVER = PatternFill("solid", fgColor="C0C0C0")
BRONZE = PatternFill("solid", fgColor="CD7F32")
ALT    = PatternFill("solid", fgColor="EBF3FB")
LGREEN = PatternFill("solid", fgColor="E2EFDA")
LRED   = PatternFill("solid", fgColor="FCE4D6")

WF = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BF = Font(name="Arial", bold=True, size=10)
RF = Font(name="Arial", size=9)
CA = Alignment(horizontal="center", vertical="center", wrap_text=True)
LA = Alignment(horizontal="left",   vertical="center")


def thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


# ─── column helpers ───────────────────────────────────────────────────────────

def col_name(year: int, kind: str) -> str:
    return f"{kind.capitalize()}_{year}"


# ─── data loading ─────────────────────────────────────────────────────────────

def load_data(path: str, use_csv_cache: bool) -> pd.DataFrame:
    csv_path = os.path.splitext(path)[0] + "_cache.csv"

    if use_csv_cache:
        if os.path.exists(csv_path):
            print(f"  Loading from CSV cache: {csv_path}")
            df = pd.read_csv(csv_path, low_memory=False)
            return df
        print("  No CSV cache found — loading xlsx and creating cache …")

    ext = os.path.splitext(path)[1].lower()
    if ext == ".csv":
        print(f"  Reading CSV: {path}")
        df = pd.read_csv(path, low_memory=False)
    else:
        print(f"  Reading Excel (this may take a moment for 112k rows) …")
        hdr_raw = pd.read_excel(path, header=None, nrows=3)
        row2 = hdr_raw.iloc[1].tolist()
        row3 = hdr_raw.iloc[2].tolist()

        cols = []
        current_year = None
        for i, (r2, r3) in enumerate(zip(row2, row3)):
            if i == 0:   cols.append("Company Name")
            elif i == 1: cols.append("NIC Code")
            elif i == 2: cols.append("Company Code")
            elif i == 3: cols.append("CIN Code")
            elif i == 4: cols.append("Sales_Products_Master")
            elif i == 5: cols.append("Sales_Services_Master")
            elif i == 6: cols.append("Total_Sales_Master")
            else:
                if r2 and str(r2).startswith("FY"):
                    current_year = int(str(r2).replace("FY", "").strip())
                kind = str(r3).strip().lower() if r3 else "unknown"
                cols.append(f"{kind.capitalize()}_{current_year}")

        raw = pd.read_excel(path, header=None, skiprows=3)
        raw.columns = cols
        df = raw.copy()

    for yr in AVAILABLE_YEARS:
        for kind in ("Products", "Services", "Total"):
            c = col_name(yr, kind)
            if c in df.columns:
                df[c] = pd.to_numeric(df[c], errors="coerce")

    df["NIC Code"] = pd.to_numeric(df["NIC Code"], errors="coerce")

    if use_csv_cache and ext != ".csv":
        df.to_csv(csv_path, index=False)
        print(f"  CSV cache saved → {csv_path}  (use --csv flag next time for instant load)")

    return df


# ─── interactive prompts ──────────────────────────────────────────────────────

def prompt_file_path() -> str:
    default = "combined_company_sales.xlsx"
    ans = input(f"\nPath to combined sales file [{default}]: ").strip()
    return ans if ans else default


def prompt_sector() -> str:
    ans = input("\n  Sector name (e.g. Steel, Pharma, Textiles): ").strip()
    return ans if ans else "Sector"


def prompt_nic_codes(df: pd.DataFrame) -> list:
    all_nics = sorted(df["NIC Code"].dropna().astype(int).unique().tolist())
    print(f"\n  {len(all_nics)} unique NIC codes found.")
    print("  Enter NIC codes separated by commas (e.g. 1, 14, 24), or Enter for ALL.")
    ans = input("  NIC code(s): ").strip()
    if not ans:
        return all_nics
    try:
        chosen = [int(x.strip()) for x in ans.split(",") if x.strip()]
    except ValueError:
        print("  ⚠  Invalid input — including all NIC codes.")
        return all_nics
    invalid = [c for c in chosen if c not in all_nics]
    if invalid:
        print(f"  ⚠  Not found, skipping: {invalid}")
    valid = [c for c in chosen if c in all_nics]
    return valid if valid else all_nics


def prompt_years() -> list:
    print(f"\n  Available years: {AVAILABLE_YEARS[0]} – {AVAILABLE_YEARS[-1]}")
    print("  Enter one of:")
    print("    • A single year          e.g.  2024")
    print("    • A range (inclusive)    e.g.  2015-2024")
    print("    • A comma list           e.g.  2018, 2020, 2022, 2024")
    while True:
        ans = input("  Year(s): ").strip()
        result = parse_years(ans)
        if result:
            return result
        print("  ⚠  Could not parse — try again.")


def parse_years(raw: str) -> list:
    """Parse '2024', '2015-2024', or '2018,2020,2022' into a sorted list of ints."""
    raw = raw.strip()
    if not raw:
        return []
    try:
        if "-" in raw and "," not in raw:
            parts = raw.split("-")
            if len(parts) == 2:
                start, end = int(parts[0].strip()), int(parts[1].strip())
                years = [y for y in range(start, end + 1) if y in AVAILABLE_YEARS]
                if years:
                    return years
        elif "," in raw:
            years = [int(x.strip()) for x in raw.split(",") if x.strip()]
            valid = [y for y in years if y in AVAILABLE_YEARS]
            invalid = [y for y in years if y not in AVAILABLE_YEARS]
            if invalid:
                print(f"  ⚠  Skipping years not in dataset: {invalid}")
            return sorted(valid)
        else:
            yr = int(raw)
            if yr in AVAILABLE_YEARS:
                return [yr]
    except ValueError:
        pass
    return []


def prompt_top_n() -> int:
    while True:
        ans = input("\n  Top N companies overall (e.g. 10): ").strip()
        try:
            n = int(ans)
            if n > 0:
                return n
            print("  ⚠  Must be a positive integer.")
        except ValueError:
            print("  ⚠  Invalid input.")


def prompt_sort_metric() -> str:
    print("\n  Rank by:")
    for k, (_, label) in SORT_CHOICES.items():
        print(f"    {k}. {label}")
    while True:
        ans = input("  Choice [1]: ").strip() or "1"
        if ans in SORT_CHOICES:
            return SORT_CHOICES[ans][0]
        print("  ⚠  Enter 1, 2, or 3.")


# ─── core filtering ───────────────────────────────────────────────────────────

def filter_and_rank(df: pd.DataFrame, nic_codes: list, year: int,
                    top_n: int, sort_metric: str) -> pd.DataFrame:
    sort_col  = col_name(year, sort_metric)
    total_col = col_name(year, "Total")

    if sort_col not in df.columns:
        return pd.DataFrame()   # year has no data — return empty

    subset = df[df["NIC Code"].isin(nic_codes)].copy()
    subset = subset[subset[total_col].notna()]

    if subset.empty:
        return pd.DataFrame()

    ranked = (
        subset
        .sort_values(sort_col, ascending=False)
        .head(top_n)
        .copy()
    )
    ranked.insert(0, "Overall Rank", range(1, len(ranked) + 1))
    return ranked


# ─── single-year ranking sheet ────────────────────────────────────────────────

def add_ranking_sheet(wb: Workbook, df_result: pd.DataFrame, year: int,
                      sort_metric: str, top_n: int, nic_codes: list, sector: str) -> None:
    """Append one ranking sheet for `year` to an existing workbook."""

    sheet_title = f"FY{year}"
    ws = wb.create_sheet(title=sheet_title)

    sort_label = SORT_CHOICES[[k for k, v in SORT_CHOICES.items() if v[0] == sort_metric][0]][1]
    nic_label  = f"{len(nic_codes)} NIC code(s): {', '.join(str(n) for n in nic_codes[:8])}" \
                 + (" …" if len(nic_codes) > 8 else "")

    display_cols = list(df_result.columns)
    ncols = len(display_cols)

    # Row 1: title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    tc = ws.cell(row=1, column=1,
                 value=f"{sector}  |  Top {top_n} Overall  |  FY {year}  |  Ranked by: {sort_label}")
    tc.font = Font(name="Arial", bold=True, color="FFFFFF", size=12)
    tc.fill = NAVY
    tc.alignment = CA
    ws.row_dimensions[1].height = 22

    # Row 2: subtitle
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    sc = ws.cell(row=2, column=1,
                 value=f"NIC codes: {nic_label}  |  Companies shown: {len(df_result)}")
    sc.font = Font(name="Arial", italic=True, size=9, color="444444")
    sc.fill = PatternFill("solid", fgColor="D6E4F7")
    sc.alignment = LA
    ws.row_dimensions[2].height = 16

    # Row 3: column headers
    col_labels = {
        "Overall Rank":          "Rank",
        "Company Name":          "Company Name",
        "NIC Code":              "NIC\nCode",
        "Company Code":          "Company\nCode",
        "CIN Code":              "CIN Code",
        "Sales_Products_Master": "Master\nProducts",
        "Sales_Services_Master": "Master\nServices",
        "Total_Sales_Master":    "Master\nTotal",
        f"Products_{year}":      f"FY{year}\nProducts",
        f"Services_{year}":      f"FY{year}\nServices",
        f"Total_{year}":         f"FY{year}\nTotal",
    }
    for ci, col in enumerate(display_cols, start=1):
        cell = ws.cell(row=3, column=ci, value=col_labels.get(col, col))
        cell.alignment = CA
        cell.border = thin_border()
        if col == f"Products_{year}":
            cell.font = WF; cell.fill = BLUE
        elif col == f"Services_{year}":
            cell.font = WF; cell.fill = GREEN
        elif col == f"Total_{year}":
            cell.font = WF; cell.fill = ORANGE
        else:
            cell.font = WF; cell.fill = NAVY
    ws.row_dimensions[3].height = 30

    # Data rows
    NUM_FMT = '#,##0.00;(#,##0.00);"-"'
    numeric_cols = {f"Products_{year}", f"Services_{year}", f"Total_{year}",
                    "Sales_Products_Master", "Sales_Services_Master", "Total_Sales_Master"}
    prev_nic     = None
    fill_toggle  = False

    for ri, (_, row) in enumerate(df_result.iterrows()):
        excel_row = ri + 4
        nic = row.get("NIC Code")
        if nic != prev_nic:
            fill_toggle = not fill_toggle
            prev_nic = nic
        row_fill = ALT if fill_toggle else None

        for ci, col in enumerate(display_cols, start=1):
            val  = row[col]
            cell = ws.cell(row=excel_row, column=ci)

            if pd.isna(val):
                cell.value = None
            elif col in numeric_cols:
                cell.value = float(val)
                cell.number_format = NUM_FMT
                cell.alignment = CA
            elif col in ("Overall Rank", "NIC Code", "Company Code", "Rank (within NIC)"):
                cell.value = int(val)
                cell.alignment = CA
            else:
                cell.value = val
                cell.alignment = LA if col == "Company Name" else CA

            cell.font   = RF
            cell.border = thin_border()
            if row_fill:
                cell.fill = row_fill

    # Column widths
    width_map = {
        "Overall Rank": 7, "Company Name": 42, "NIC Code": 8,
        "Company Code": 12, "CIN Code": 24,
        "Sales_Products_Master": 14, "Sales_Services_Master": 14, "Total_Sales_Master": 14,
        f"Products_{year}": 16, f"Services_{year}": 16, f"Total_{year}": 16,
    }
    for ci, col in enumerate(display_cols, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = width_map.get(col, 14)

    ws.freeze_panes = "C4"


# ─── cross-year summary sheet ─────────────────────────────────────────────────

def add_summary_sheet(wb: Workbook, results: dict, years: list,
                      sort_metric: str, top_n: int, sector: str) -> None:
    """
    Add a 'Summary' sheet showing top-N companies as rows and years as columns,
    with their total sales for quick cross-year comparison.
    Also highlights rank changes (up/down) year-over-year.
    """
    ws = wb.create_sheet(title="Summary – All Years")

    sort_label = SORT_CHOICES[[k for k, v in SORT_CHOICES.items() if v[0] == sort_metric][0]][1]

    # Collect union of all companies that appeared in any year
    all_companies = {}   # (company, nic) -> {year: total_sales}
    for yr, df in results.items():
        if df.empty:
            continue
        for _, row in df.iterrows():
            key = (row["Company Name"], int(row["NIC Code"]))
            if key not in all_companies:
                all_companies[key] = {}
            tc = col_name(yr, "Total")
            all_companies[key][yr] = row.get(tc, np.nan)

    if not all_companies:
        return

    # Build summary dataframe
    rows = []
    for (name, nic), yr_data in all_companies.items():
        entry = {"Company Name": name, "NIC Code": nic}
        for yr in years:
            entry[f"Total_{yr}"] = yr_data.get(yr, np.nan)
        rows.append(entry)

    df_sum = pd.DataFrame(rows).sort_values(["NIC Code", f"Total_{years[-1]}"],
                                             ascending=[True, False], na_position="last")

    # ── Header rows ──
    ncols = 2 + len(years)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    tc = ws.cell(row=1, column=1,
                 value=f"{sector}  |  Cross-Year Sales Summary  |  Top {top_n} Overall  |  Metric: {sort_label}")
    tc.font = Font(name="Arial", bold=True, color="FFFFFF", size=12)
    tc.fill = NAVY
    tc.alignment = CA
    ws.row_dimensions[1].height = 22

    # Column headers
    hdrs = ["Company Name", "NIC Code"] + [f"FY {yr}\nTotal (Rs. Cr.)" for yr in years]
    for ci, h in enumerate(hdrs, start=1):
        cell = ws.cell(row=2, column=ci, value=h)
        cell.font = WF
        cell.fill = NAVY if ci <= 2 else BLUE
        cell.alignment = CA
        cell.border = thin_border()
    ws.row_dimensions[2].height = 30

    # Data
    NUM_FMT = '#,##0.00;(#,##0.00);"-"'
    prev_nic    = None
    fill_toggle = False

    for ri, (_, row) in enumerate(df_sum.iterrows()):
        excel_row = ri + 3
        nic = row["NIC Code"]
        if nic != prev_nic:
            fill_toggle = not fill_toggle
            prev_nic = nic
        row_fill = ALT if fill_toggle else None

        # Company name
        c1 = ws.cell(row=excel_row, column=1, value=row["Company Name"])
        c1.font = RF; c1.alignment = LA; c1.border = thin_border()
        if row_fill: c1.fill = row_fill

        # NIC code
        c2 = ws.cell(row=excel_row, column=2, value=int(nic))
        c2.font = RF; c2.alignment = CA; c2.border = thin_border()
        if row_fill: c2.fill = row_fill

        # Year totals with YoY highlight
        prev_val = None
        for yi, yr in enumerate(years):
            val  = row.get(f"Total_{yr}", np.nan)
            cell = ws.cell(row=excel_row, column=3 + yi)
            cell.font = RF
            cell.border = thin_border()
            cell.alignment = CA

            if pd.isna(val):
                cell.value = None
                cell.fill = row_fill or PatternFill()
            else:
                cell.value = float(val)
                cell.number_format = NUM_FMT
                # YoY colour: green if up, red if down vs previous year
                if prev_val is not None and not pd.isna(prev_val):
                    if val > prev_val:
                        cell.fill = LGREEN
                    elif val < prev_val:
                        cell.fill = LRED
                    else:
                        cell.fill = row_fill or PatternFill()
                else:
                    cell.fill = row_fill or PatternFill()
            prev_val = val if not pd.isna(val) else prev_val

    # Column widths
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 9
    for i in range(len(years)):
        ws.column_dimensions[get_column_letter(3 + i)].width = 16

    ws.freeze_panes = "C3"

    # ── Legend ──
    legend_row = len(df_sum) + 5
    ws.merge_cells(start_row=legend_row, start_column=1, end_row=legend_row, end_column=4)
    lg = ws.cell(row=legend_row, column=1,
                 value="Legend:  Green = sales higher than previous year    Red = sales lower than previous year    Blank = no data for that year")
    lg.font = Font(name="Arial", italic=True, size=8, color="666666")
    lg.alignment = LA


# ─── NIC summary sheet ────────────────────────────────────────────────────────

def add_nic_summary_sheet(wb: Workbook, results: dict, years: list,
                          top_n: int, sector: str) -> None:
    """One sheet showing aggregate totals per NIC code across all years."""
    ws = wb.create_sheet(title="NIC Summary")

    ncols = 1 + len(years) * 2   # NIC + (TotalSales, AvgSales) per year
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    tc = ws.cell(row=1, column=1,
                 value=f"{sector}  |  NIC Aggregate Sales (Rs. Cr.)  |  Top {top_n} Overall")
    tc.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    tc.fill = NAVY; tc.alignment = CA
    ws.row_dimensions[1].height = 20

    # Header row 2: NIC | FY2014 Total | FY2014 Avg | FY2015 ...
    ws.cell(row=2, column=1, value="NIC Code").font = WF
    ws.cell(row=2, column=1).fill = NAVY
    ws.cell(row=2, column=1).alignment = CA
    ws.cell(row=2, column=1).border = thin_border()

    for yi, yr in enumerate(years):
        c_tot = 2 + yi * 2
        c_avg = c_tot + 1
        for ci, lbl in [(c_tot, f"FY{yr}\nTotal"), (c_avg, f"FY{yr}\nAvg")]:
            cell = ws.cell(row=2, column=ci, value=lbl)
            cell.font = WF; cell.fill = BLUE; cell.alignment = CA; cell.border = thin_border()
    ws.row_dimensions[2].height = 28

    # Collect NIC codes
    all_nics = sorted(set(
        int(nic)
        for df in results.values() if not df.empty
        for nic in df["NIC Code"].dropna()
    ))

    NUM_FMT = '#,##0.00'
    for ri, nic in enumerate(all_nics):
        r = ri + 3
        fill = ALT if ri % 2 == 0 else None
        c1 = ws.cell(row=r, column=1, value=nic)
        c1.font = BF; c1.alignment = CA; c1.border = thin_border()
        if fill: c1.fill = fill

        for yi, yr in enumerate(years):
            df = results.get(yr, pd.DataFrame())
            tc_col = col_name(yr, "Total")
            if df.empty or tc_col not in df.columns:
                tot, avg = None, None
            else:
                grp = df[df["NIC Code"] == nic][tc_col].dropna()
                tot = round(float(grp.sum()), 2) if len(grp) else None
                avg = round(float(grp.mean()), 2) if len(grp) else None

            for ci, val in [(2 + yi * 2, tot), (3 + yi * 2, avg)]:
                cell = ws.cell(row=r, column=ci, value=val)
                cell.font = RF; cell.alignment = CA; cell.border = thin_border()
                if val is not None:
                    cell.number_format = NUM_FMT
                if fill:
                    cell.fill = fill

    ws.column_dimensions["A"].width = 10
    for i in range(len(years)):
        ws.column_dimensions[get_column_letter(2 + i * 2)].width = 15
        ws.column_dimensions[get_column_letter(3 + i * 2)].width = 15
    ws.freeze_panes = "B3"


# ─── main ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Filter top-N companies by NIC code across years.")
    parser.add_argument("--file",    default="", help="Path to input xlsx/csv file")
    parser.add_argument("--csv",     action="store_true", help="Use/create CSV cache for faster loading")
    parser.add_argument("--sector",  default="", help="Sector name (e.g. Steel, Pharma)")
    parser.add_argument("--nic",     default="", help="Comma-separated NIC codes (blank = all)")
    parser.add_argument("--years",   default="", help="Year(s): single '2024', range '2015-2024', or list '2018,2020,2024'")
    parser.add_argument("--top",     type=int, default=0, help="Top N companies per NIC code")
    parser.add_argument("--metric",  default="", help="Sort metric: total / products / services")
    parser.add_argument("--output",  default="", help="Output xlsx path (auto-generated if omitted)")
    args = parser.parse_args()

    print("\n" + "═"*62)
    print("   TOP-N COMPANY FILTER  |  MULTI-YEAR")
    print("═"*62)

    # ── File ──
    file_path = args.file or prompt_file_path()
    if not os.path.exists(file_path):
        sys.exit(f"File not found: {file_path}")

    # ── Load ──
    print("\n[1/6] Loading data …")
    t0 = time.time()
    df = load_data(file_path, use_csv_cache=args.csv)
    print(f"  Loaded {len(df):,} rows in {time.time()-t0:.1f}s")

    # ── Sector ──
    sector = args.sector.strip() if args.sector.strip() else prompt_sector()
    print(f"  Sector: {sector}")

    # ── NIC codes ──
    print("\n[2/6] NIC code selection")
    if args.nic:
        all_nics = sorted(df["NIC Code"].dropna().astype(int).unique().tolist())
        chosen_nics = [int(x.strip()) for x in args.nic.split(",") if x.strip()]
        invalid = [c for c in chosen_nics if c not in all_nics]
        if invalid:
            print(f"  ⚠  Not found, skipping: {invalid}")
        chosen_nics = [c for c in chosen_nics if c in all_nics]
    else:
        chosen_nics = prompt_nic_codes(df)
    print(f"  Selected: {chosen_nics[:10]}{'…' if len(chosen_nics)>10 else ''} ({len(chosen_nics)} code(s))")

    # ── Years ──
    print("\n[3/6] Year selection")
    if args.years:
        chosen_years = parse_years(args.years)
        if not chosen_years:
            sys.exit(f"Could not parse --years '{args.years}'. Use: 2024 | 2015-2024 | 2018,2020,2024")
    else:
        chosen_years = prompt_years()
    yr_range_label = f"FY{chosen_years[0]}" if len(chosen_years) == 1 \
                     else f"FY{chosen_years[0]}–FY{chosen_years[-1]}"
    print(f"  Years: {chosen_years}  ({len(chosen_years)} year(s))")

    # ── Top N & metric ──
    print("\n[4/6] Top-N and sort metric")
    top_n = args.top if args.top > 0 else prompt_top_n()
    metric = args.metric.lower() if args.metric.lower() in ("total", "products", "services") \
             else prompt_sort_metric()
    print(f"  Top {top_n} | Sorted by: {metric}")

    # ── Filter each year ──
    print("\n[5/6] Filtering and ranking …")
    results = {}
    for yr in chosen_years:
        res = filter_and_rank(df, chosen_nics, yr, top_n, metric)
        results[yr] = res
        status = f"{len(res)} companies" if not res.empty else "no data"
        print(f"  FY{yr}: {status}")

    non_empty = {yr: r for yr, r in results.items() if not r.empty}
    if not non_empty:
        print("  ⚠  No data found for any selected year/NIC combination.")
        sys.exit(0)

    # ── Build workbook ──
    print("\n[6/6] Writing Excel …")
    wb = Workbook()
    wb.remove(wb.active)   # remove default blank sheet

    for yr in chosen_years:
        if not results[yr].empty:
            add_ranking_sheet(wb, results[yr], yr, metric, top_n, chosen_nics, sector)
            print(f"  ✓ Sheet FY{yr}")
        else:
            print(f"  – FY{yr} skipped (no data)")

    add_summary_sheet(wb, results, [y for y in chosen_years if not results[y].empty],
                      metric, top_n, sector)
    print("  ✓ Sheet: Summary – All Years")

    add_nic_summary_sheet(wb, results, [y for y in chosen_years if not results[y].empty],
                          top_n, sector)
    print("  ✓ Sheet: NIC Summary")

    # ── Output path ──
    if args.output:
        out_path = args.output
    else:
        safe_sector = sector.replace(" ", "_").replace("/", "-")
        auto_name   = f"{safe_sector}_{yr_range_label}_top{top_n}.xlsx"
        ans = input(f"\n  Output file name [{auto_name}]: ").strip()
        out_path = ans if ans else auto_name

    wb.save(out_path)
    print(f"\n✓ Saved → {out_path}")
    print(f"  Sheets: {[s.title for s in wb.worksheets]}")
    print("═"*62 + "\n")


if __name__ == "__main__":
    main()
